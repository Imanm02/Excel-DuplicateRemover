import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

def apply_styles_to_excel(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    font = Font(name='Vazirmatn')
    alignment = Alignment(horizontal='center', vertical='center')
    light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row in sheet:
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            cell.number_format = '@'
            if cell.row == 1:
                cell.fill = light_yellow_fill
            else:
                cell.fill = light_green_fill

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    wb.save(filename)

all_sheets = pd.read_excel("input.xlsx", sheet_name=None, dtype=str)

full_df = pd.concat(all_sheets.values(), ignore_index=True)

unique_df = full_df.drop_duplicates(subset=['Name', 'Surname'])

output_filename = "output.xlsx"
unique_df.to_excel(output_filename, index=False)

apply_styles_to_excel(output_filename)
